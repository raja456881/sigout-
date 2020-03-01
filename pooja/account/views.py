from django.shortcuts import render , redirect
from django.contrib.auth.models import User
from django.contrib import auth
from .models import Hotel
from django.core.files.storage import FileSystemStorage
import openpyxl



# Create your views here.
def home(request):
    return render(request , 'account/index.html')
def login(request):
    if request.method=='POST':
        uname=request.POST['username']
        passs=request.POST['pass']
        user=auth.authenticate(username=uname , password=passs)
        if user is not None:
            auth.login(request , user)
            return render(request, 'account/login.html')
        else:
            return render(request , 'account/index.html' , {'error':"Invaild login "})

    else:
        return render(request , "account/index.html")



def res(request):
    if request.method=='POST':
        if request.POST['password']==request.POST['password confirm']:
            try:
                user=User.objects.get(username=request.POST['username'])
                return render(request, 'account/resg.html', {'error': "User Already  been Taken"})


            except User.DoesNotExist:
                user=User.objects.create_user(username=request.POST['username'], password=request.POST['password'])
                auth.login(request , user)
                return redirect(home)

        else:
            return render( request,'account/resg.html' , {'error':"Password Don't Match"})
    else:
        return render(request, 'account/resg.html')


def logout(request):
    auth.logout(request)
    return redirect(home)


def image(request):
    content={}
    if request.method=='POST':
        p=request.FILES['image']
        fs = FileSystemStorage()
        name = fs.save(p.name, p)
        content['url']=fs.url(name)
    return render(request ,'account/ra.html' , content)



def excel(request):
    if "GET" == request.method:
        return render(request, 'account/excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, 'account/excel.html', {"excel_data":excel_data})
